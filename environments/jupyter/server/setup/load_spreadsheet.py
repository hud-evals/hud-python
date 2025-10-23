"""Load specific spreadsheet tasks from SpreadsheetBench dataset."""

import json
import os
from pathlib import Path
from ..config import VOLUMES_PATH
from ..tools import JupyterToolWithRecord
from hud.server import MCPRouter

router = MCPRouter()


@router.tool("load_spreadsheet")
async def load_spreadsheet(id: str, dataset_path: str = "all_data_912"):
    """
    Load spreadsheet for a SpreadsheetBench task (single-turn approach).

    Loads only the first instance (1_) following SpreadsheetBench evaluation pattern:
    1. Agent solves the task on instance 1
    2. Solution is applied to instances 2 & 3 for generalization testing

    Args:
        id: The task ID (e.g., "59196", "99-24")

    Returns:
        Task information with file preview
    """
    # Paths inside container
    dataset_path = os.path.join(VOLUMES_PATH, dataset_path)
    dataset_json = os.path.join(dataset_path, "dataset.json")
    spreadsheet_dir = os.path.join(dataset_path, "spreadsheet", id)

    # Load dataset and find task
    if not Path(dataset_json).exists():
        return f"❌ Dataset not found at {dataset_json}"

    with open(dataset_json) as f:
        dataset = json.load(f)

    task_info = next((t for t in dataset if str(t["id"]) == str(id)), None)
    if not task_info:
        return f"❌ Task {id} not found in dataset"

    if not Path(spreadsheet_dir).exists():
        return f"❌ Spreadsheet directory not found: {spreadsheet_dir}"

    # Connect to the shared kernel
    jupyter_tool = JupyterToolWithRecord.from_shared_kernel("SpreadSheetBench")

    # Load data - simple and literal for KMP generalization
    code = f"""
import openpyxl
wb = openpyxl.load_workbook("{spreadsheet_dir}/1_{id}_input.xlsx")
ws = wb.active

print(f"✅ Loaded: 1_{id}_input.xlsx")
print(f"Variables: wb (workbook), ws (worksheet)")
print(f"Dimensions: {{ws.max_row}} rows x {{ws.max_column}} columns")
print(f"\\nFirst 5 rows:")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
    print(f"Row {{i}}: {{row}}")
"""

    # Execute code (automatically recorded if successful)
    result = await jupyter_tool._execute(code)

    # Return formatted response
    return f"""✅ Loaded SpreadsheetBench Task: {id}

{result}

💡 File Paths:
   Input:  {spreadsheet_dir}/1_{id}_input.xlsx
   Output: {spreadsheet_dir}/1_{id}_output.xlsx

⚠️  Note: Data is already loaded into workbook variable `wb` shown above. Use them directly! 
⚠️  Please use openyxl instead of pandas, to keep the original xlxs shape!
⚠️  Solve using instance 1. Solution will be applied to instances 2 & 3 for testing.
"""
